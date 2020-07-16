#!/usr/bin/python3
"""
   test/demo program that extract comments from pdf files into a Excel
   command line:
   %s [-d] [-o output.xls] input0.pdf [input1.pdf .. inputN.pdf]
   -xls: create an xls(using xlwt) rather than xlsx(using openxl) file
   -d: open Excel output at the end of extraction
   -o: prode the output Excel name/path ; if not present the file is created
       in temp folder named "comments on **PDFfile**.xlsx"
       if
    if no parameters (mainly for idle test), the pdf filename is asked for
"""
from collections import OrderedDict
from datetime import datetime
import locale
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except: #pylint: disable=bare-except
    print("openpyxl not loaded", file=sys.stderr)
try:
    import xlwt
except: #pylint: disable=bare-except
    print("xlwt not loaded", file=sys.stderr)

import pypdf as PDF


locale.setlocale(locale.LC_ALL, locale.getdefaultlocale()[0])



def list_outlines(pdf_s, outl=None):
    """
    provide as a list of the outlines as tuple Title,Page(0 based),Vertical position in %
    """
    if outl is None:
        lst_ = [
            ("-", 0, 0),
        ]
        outl = pdf_s.getOutlines()
    else:
        lst_ = []
    if isinstance(outl, list):
        for k__ in outl:
            lst_ += list_outlines(pdf_s, k__)
    else:
        try:
            top = outl["/Top"]
        except: #pylint: disable=bare-except
            top = 0
        try:
            pp_ = pdf_s.MyPages[outl.page.idnum]
            lst_.append((outl.title, pp_[0], 100.0 * (1.0 - float(top / pp_[1]))))
        except: #pylint: disable=bare-except
            print("trouble with page idnum", outl.page.idnum)
    return lst_


def list_annots(pdf_s):
    """
    provide as a list of the comments with the response saved in .irt_str field
    the list is indexed with idnums
    """
    lst_ = OrderedDict()
    for pn_ in range(pdf_s.numPages):
        p__ = pdf_s.getPage(pn_)
        try:
            a__ = p__.get("/Annots").getObject()
            if not isinstance(a__, list):
                a__ = [a__]
            for b__ in a__:
                o__ = b__.getObject()
                if o__["/Subtype"] == "/Text":
                    try:
                        o__["/P"]  # le champs '/P' etant optionnel on le reconstruit...
                    except: #pylint: disable=bare-except
                        o__.update({PDF.NameObject("/P"): p__.indirectRef})
                    o__.irt = {}
                    lst_[b__.idnum] = o__
        except: #pylint: disable=bare-except
            pass
    # copy the information into the original comment
    for o__ in lst_.values():
        if "/IRT" in o__:
            t__ = o__["/Contents"]
            if isinstance(t__, bytes):
                t__ = t__.replace(b"\r", b"\n").decode("unicode_escape")
            lst_[o__.rawGet("/IRT").idnum].irt[o__["/M"]] = "%s (%s):\n%s" % (
                o__["/T"],
                datetime.strptime(o__["/M"][2:10], "%Y%m%d").strftime("%x"),
                t__,
            )
    # concat all replied comments into one string to ease insertion later...
    for o__ in lst_.values():
        o__.irt_str = "\n".join([o__.irt[x] for x in sorted(o__.irt.keys())])
    return lst_


def find_outline(outl, pa_, pe_):
    """
    provide the outline just above the position (of the comment)
    """
    m__ = None
    for o__ in outl:
        if (o__[1] < pa_) or ((o__[1] == pa_) and (o__[2] <= pe_)):
            m__ = o__
    return m__

def app_ws(sht, tup):
    """
    append a new line to sheet iaw selected library
    """
    if xlsx_out:
        sht.append(tup)
    else:
        sty1 = xlwt.easyxf("alignment: vertical Top")
        sty = xlwt.easyxf("alignment: wrap True, vertical Top")
        rw_ = sht.row(sht.active_row)
        sht.active_row += 1
        i__ = 0
        for v__ in tup:
            if i__ == 0:
                rw_.write(i__, v__, sty1)
            else:
                rw_.write(i__, v__, sty)
            i__ += 1

def prepare_workbook():
    """
        prepare an empty destination workbook
    """
    colwidth = enumerate((5, 5, 5, 25, 15, 90, 90))
    if xlsx_out:
        wb_ = Workbook()
        ws_ = wb_.active
        for i__, cw_ in colwidth:
            ws_.column_dimensions[get_column_letter(i__+ 1)].width = cw_
    else:
        wb_ = xlwt.Workbook()
        ws_ = wb_.add_sheet("Comments 1")
        for i__, cw_ in colwidth:
            ws_.col(i__).width = cw_ * 256
        ws_.active_row = 0
    app_ws(ws_, ("Doc", "Page", "Pos", "Chapt", "Originator", "Comment", "Answer"))
    return wb_, ws_

def pdf_to_excel(pdf_source, ws_):
    """
        extract comments to excel
    """
    doc_ = os.path.splitext(os.path.basename(os.path.abspath(pdf_source.filepath)))[0]
    # check if decryption is required
    if pdf_source.isEncrypted:
        pdf_source.decrypt("")

    # MyPages will store the matching table page.idnum => pagenumer,page_height
    pdf_source.MyPages = {}

    for i__, p__ in enumerate(pdf_source.pages):
        pdf_source.MyPages[p__.indirectRef.idnum] = [i__, p__["/MediaBox"][3]]

    # extract the list of OutLines into MyOutlines
    pdf_source.MyOutlines = list_outlines(pdf_source)

    # extract the comments into MyAnnots
    pdf_source.MyAnnots = list_annots(pdf_source)


    # sort the comments in the order (Page, vertical position, date)
    lst = {}
    for p__ in pdf_source.MyAnnots.values():
        pp_ = pdf_source.MyPages[p__.rawGet("/P").idnum]
        pc_ = 100.0 * (1.0 - float(int(p__["/Rect"][1]) / pp_[1]))
        lst[(pp_[0], pc_, p__["/M"])] = p__

    # fill the xl sheet with the comments
    for x__ in sorted(lst.keys()):
        p__ = lst[x__]
        if "/IRT" in p__:
            # the comments with IRT are already present in the original comment irt field,
            # we can ignore this one
            continue

        auth = p__["/T"]
        if isinstance(auth, bytes):
            auth = auth.decode("unicode_escape")
        try:
            cont = p__["/Contents"]
        except: #pylint: disable=bare-except
            cont = ""
        if isinstance(cont, bytes):
            cont = cont.replace(b"\r", b"\n").decode("unicode_escape")
        if isinstance(p__.irt_str, bytes):
            p__.irt_str = p__.irt_str.replace(b"\r", b"\n").decode("unicode_escape")

        app_ws(ws_,
               (
                   doc_,
                   pdf_source.getPageLabel(x__[0]),
                   "%.0f %%" % x__[1],
                   find_outline(pdf_source.MyOutlines, x__[0], x__[1])[0],
                   auth,
                   cont,
                   p__.irt_str,
               )
              )
    return ws_

######################### main ########################

if sys.argv[0].upper().find("PYTHON.EXE") >= 0:
    del sys.argv[0]

if len(sys.argv) == 1:
    print(globals()["__doc__"]%(sys.argv[0]))
    sys.argv.append(input("pdf file to scan:"))

del sys.argv[0] # remove also call program

xlsx_out = True
for i_ in range(len(sys.argv)):
    if sys.argv[i_].upper() == '-XLS':
        xlsx_out = False
        del sys.argv[i_]
        break

start_excel = ("idlelib.run" in sys.modules)
for i_ in range(len(sys.argv)):
    if sys.argv[i_].upper() == '-D':
        start_excel = True
        del sys.argv[i_]
        break

if "-o" in sys.argv:
    i_ = sys.argv.index("-o")
    xlFile = os.path.expandvars(sys.argv[i_ + 1])
    del sys.argv[i_] # -o
    del sys.argv[i_] # xl_file
else:
    tempFolder = os.environ["TEMP"].replace("\\", "/")
    if tempFolder[-1] != "/":
        tempFolder += "/"
    xlFile = None

wb, ws = prepare_workbook() # prepare a first wb and workheet
for fn in sys.argv: #the remaining command line only contains pdf to process
    ws = pdf_to_excel(PDF.PdfFileReader(fn), ws)
    # post insertion formating
    if xlsx_out:
        for row in ws.iter_rows():
            for cell in row:
                if cell.column > 1:
                    cell.alignment = cell.alignment.copy(wrapText=True, vertical="top")
    #else (!xlsx_out) alignment is done during writing
    # save the file
    if xlFile is None:
        doc = os.path.splitext(os.path.basename(os.path.abspath(fn)))[0]
        xf = tempFolder + "Comments on " + doc + (".xlsx" if xlsx_out else ".xls")
        wb.save(xf)
        wb, ws = prepare_workbook() # we prepare a new empty workbook for next pdf file
        print(xf)
        if start_excel:
            os.startfile(xf)
if xlFile:
    xf = xlFile
    wb.save(xf)
    if start_excel:
        os.startfile(xf)
